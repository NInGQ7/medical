# -*- coding: utf-8 -*-
"""
医疗设备数据融合主程序
"""

import os
import sys
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 添加项目根目录到路径
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from src.fusion_engine import FusionEngine
from src.parameter_preprocessor import ParameterPreprocessor
from config.config import CONFIG


class MedicalDeviceFusion:
    """医疗设备数据融合处理器"""
    
    def __init__(self, config: dict = None):
        """
        初始化
        
        Args:
            config: 配置字典，如果为None则使用默认配置
        """
        self.config = config if config else CONFIG
        self.fusion_engine = FusionEngine()
        self.preprocessor = ParameterPreprocessor()
        self.log_messages = []
    
    def log(self, message: str):
        """记录日志"""
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_msg = f"[{timestamp}] {message}"
        self.log_messages.append(log_msg)
        print(log_msg)
    
    def process_excel(self, input_file: str, output_file: str = None) -> bool:
        """
        处理Excel文件
        
        Args:
            input_file: 输入Excel文件路径
            output_file: 输出Excel文件路径，如果为None则自动生成
            
        Returns:
            是否处理成功
        """
        try:
            # 检查文件是否存在
            if not os.path.exists(input_file):
                self.log(f"错误: 文件不存在 - {input_file}")
                return False
            
            self.log(f"开始处理文件: {input_file}")
            
            # 读取Excel文件
            df = pd.read_excel(input_file)
            self.log(f"成功读取Excel文件，共 {len(df)} 行数据")
            
            # 检查数据格式
            if df.shape[1] < 2:
                self.log("错误: Excel文件至少需要2列（参数名称 + 至少1个供应商数据）")
                return False
            
            self.log(f"检测到 {df.shape[1] - 1} 个供应商列")
            
            # 准备结果列
            fused_data = []  # 每行的融合结果
            fusion_types = []  # 每行的融合类型
            all_fused_results = []  # 所有融合结果的列表
            
            # 【新增】供应商达标判定颜色
            num_vendors = df.shape[1] - 1
            vendor_compliance_colors = [[] for _ in range(num_vendors)]
            
            # 逐行处理
            total_rows = len(df)
            for index, row in df.iterrows():
                param_name = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                vendor_values = row.iloc[1:].tolist()
                
                # 第一步：参数预处理 - 拆分整合参数
                # 关键：当供应商数据是整合参数时（包含多个类型），根据当前行的参数名称提取相关数据
                preprocessed_values = []
                for value in vendor_values:
                    if pd.notna(value):
                        value_str = str(value).strip()
                        # 检查是否为整合参数（包含2个以上的分隔符：逗号、顿号、分号）
                        if self.preprocessor.is_integrated_params(value_str):
                            # 是整合参数，需要根据参数名称拆分提取
                            preprocessed = self.preprocessor.preprocess(value_str, param_name)
                            preprocessed_values.append(preprocessed)
                        else:
                            # 不是整合参数，直接使用
                            preprocessed_values.append(value_str)
                    else:
                        preprocessed_values.append(value)
                
                # 第二步：融合（使用预处理后的数据）
                result, f_type = self.fusion_engine.process_row(param_name, preprocessed_values)
                fused_data.append(result)
                fusion_types.append(f_type)
                all_fused_results.append(f"【{param_name}】{result}")
                
                # 【修改】第三步：供应商达标判定（需人工审核的不判断）
                if "需人工审核" in f_type:
                    # 需人工审核的行，不判断供应商是否达标，不添加颜色
                    for vendor_idx in range(len(preprocessed_values)):
                        vendor_compliance_colors[vendor_idx].append('none')
                else:
                    # 非人工审核的行，进行达标判定
                    for vendor_idx, preprocessed_val in enumerate(preprocessed_values):
                        color = self._evaluate_supplier_compliance(
                            preprocessed_val, result, param_name, f_type
                        )
                        vendor_compliance_colors[vendor_idx].append(color)
                
                # 进度显示
                if (index + 1) % 10 == 0 or (index + 1) == total_rows:
                    progress = (index + 1) / total_rows * 100
                    self.log(f"处理进度: {index + 1}/{total_rows} ({progress:.1f}%)")
            
            # 添加融合结果列
            df['融合数据'] = fused_data
            df['融合类型'] = fusion_types
            
            # 添加合并列：将所有参数融合结果合并为一个单元格
            merged_content = "\n".join(all_fused_results)
            df['合并数据'] = merged_content
            
            # 生成输出文件名
            if not output_file:
                input_path = Path(input_file)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_file = str(input_path.parent / f"{input_path.stem}_融合结果_{timestamp}.xlsx")
            
            # 保存结果（空值保持原样，不做任何处理）
            df.to_excel(output_file, index=False, engine='openpyxl')
            
            # 【新增】在第一行添加说明
            self._add_header_instruction(output_file, df.shape[1])
            
            # 为"需人工审核"的融合数据行添加黄色背景
            self._apply_yellow_highlight_for_manual_review(output_file, fusion_types)
            
            # 【新增】为供应商列添加达标颁色
            self._apply_vendor_compliance_colors(output_file, vendor_compliance_colors, total_rows)
            
            # 合并"合并数据"列的所有单元格
            self._merge_merged_data_column(output_file, len(df))
            
            # 设置所有单元格自动换行，合并列顶端对齐
            self._format_cells(output_file, len(df))
            
            self.log(f"融合结果已保存到: {output_file}")
            
            # 输出统计信息
            if self.config.get('output_statistics', True):
                self._output_statistics()
            
            # 保存日志
            if self.config.get('output_log', True):
                self._save_log()
            
            return True
            
        except Exception as e:
            self.log(f"处理过程中发生错误: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            return False
    
    def _output_statistics(self):
        """输出统计信息"""
        stats = self.fusion_engine.get_statistics()
        
        self.log("\n" + "="*50)
        self.log("融合统计信息:")
        self.log("="*50)
        
        for fusion_type, info in stats.items():
            count = info['count']
            percent = info['percent']
            self.log(f"{fusion_type}: {count} 条 ({percent}%)")
        
        self.log("="*50 + "\n")
    
    def _evaluate_supplier_compliance(self, supplier_value: str, fused_value: str, 
                                      parameter_name: str, fusion_type: str) -> str:
        """
        【修改】评估供应商数据是否满足融合要求
        注意：不对不满足的单元格进行红色标记，仅标记达标（蓝色）和无数据（灰色）
        
        Args:
            supplier_value: 供应商数据
            fused_value: 融合后的数据
            parameter_name: 参数名称
            fusion_type: 融合类型
            
        Returns:
            颜色: 'blue'(达标), 'gray'(无数据), 'none'(不标记)
        """
        import pandas as pd
        from config.config import PARAM_RULES
        
        # 空值判断
        if pd.isna(supplier_value) or not str(supplier_value).strip():
            return 'gray'
        
        supplier_str = str(supplier_value).strip()
        fused_str = str(fused_value).strip()
        
        # 【增强1】如果供应商数据包含融合参数，则满足
        if fused_str.lower() in supplier_str.lower():
            return 'blue'
        
        # 获取参数规则
        rule = self.fusion_engine.get_rule_for_parameter(parameter_name)
        rule_type = rule.get('type', 'auto')
        
        # 文本类参数
        if rule_type == 'text' or '相似度' in fusion_type or '语义' in fusion_type:
            # 【增强2】检查语义等价
            if self._check_semantic_equivalent(supplier_str, fused_str):
                return 'blue'
            
            similarity = self.fusion_engine.text_processor.calculate_similarity(
                supplier_str, fused_str, 'token_set'
            )
            # 【增强3】相似度判断：相似度≥60%且覆盖关键内容时标记为满足（蓝色）
            threshold = 60  # 改为 60% 相似度
            
            if similarity >= threshold:
                # 【增强】检查是否覆盖关键内容（不强制要求）
                # 简化逻辑：只要相似度≥60%就认为满足
                return 'blue'
            
            return 'none'
        
        # 数字类参数
        if rule_type == 'numeric' or '范围' in fusion_type or '单位转换' in fusion_type:
            return self._evaluate_numeric_compliance(supplier_str, fused_str, rule)
        
        # 多值类参数
        if rule_type == 'multi_value':
            return self._evaluate_multi_value_compliance(supplier_str, fused_str, rule)
        
        # 默认：精确匹配
        if supplier_str.lower() == fused_str.lower():
            return 'blue'
        
        # 【修改】不匹配返回'none'而非'red'
        return 'none'
    
    def _evaluate_numeric_compliance(self, supplier_str: str, fused_str: str, rule: dict) -> str:
        """数字类参数达标判定，达标标蓝色、不达标不标色"""
        import re
        
        # 提取数字
        supplier_nums = self.fusion_engine.numeric_processor.extract_numeric_info(supplier_str)
        fused_nums = self.fusion_engine.numeric_processor.extract_numeric_info(fused_str)
        
        if not supplier_nums or not fused_nums:
            return 'gray'
        
        supplier_val = supplier_nums[0]['value']
        tolerance = rule.get('tolerance', self.config.get('numeric_tolerance', 0.05))
        
        # 【增强3】检查融合参数是否包含比较符号（≥1、≤15等）
        comparison_match = re.search(r'[≥≤><=＞＜]+\s*(\d+\.?\d*)', fused_str)
        if comparison_match:
            operator = re.search(r'[≥≤><=＞＜]+', fused_str).group()
            threshold_val = float(comparison_match.group(1))
            
            # 判断供应商数值是否满足比较条件
            # 【修改】达标返回'blue'，不达标返回'none'
            if '≥' in operator or '>=' in operator or '＞=' in operator:
                return 'blue' if supplier_val >= threshold_val else 'none'
            elif '≤' in operator or '<=' in operator or '＜=' in operator:
                return 'blue' if supplier_val <= threshold_val else 'none'
            elif '>' in operator or '＞' in operator:
                return 'blue' if supplier_val > threshold_val else 'none'
            elif '<' in operator or '＜' in operator:
                return 'blue' if supplier_val < threshold_val else 'none'
            elif '=' in operator:
                error = abs(supplier_val - threshold_val) / threshold_val if threshold_val != 0 else 0
                return 'blue' if error <= tolerance else 'none'
        
        # 判断是否为范围
        if '-' in fused_str and len(fused_nums) >= 2:
            # 范围判断
            min_val = min(n['value'] for n in fused_nums)
            max_val = max(n['value'] for n in fused_nums)
            # 【修改】达标返回'blue'，不达标返回'none'
            return 'blue' if min_val <= supplier_val <= max_val else 'none'
        else:
            # 单值判断（允许误差）
            fused_val = fused_nums[0]['value']
            error = abs(supplier_val - fused_val) / fused_val if fused_val != 0 else 0
            # 【修改】达标返回'blue'，不达标返回'none'
            return 'blue' if error <= tolerance else 'none'
    
    def _evaluate_multi_value_compliance(self, supplier_str: str, fused_str: str, rule: dict) -> str:
        """多值类参数达标判定，达标标蓝色、不达标不标色"""
        separator = rule.get('separator', '×')
        supplier_dict = self.fusion_engine.numeric_processor.parse_multi_value(supplier_str, separator)
        fused_dict = self.fusion_engine.numeric_processor.parse_multi_value(fused_str, separator)
        
        if not supplier_dict or not fused_dict:
            return 'gray'
        
        # 检查每个维度是否达标
        for key, required_val in fused_dict.items():
            supplier_val = supplier_dict.get(key, 0)
            if supplier_val < required_val:
                # 【修改】不达标返回'none'
                return 'none'
        
        # 所有维度都满足，返回'blue'
        return 'blue'
    
    def _check_semantic_equivalent(self, text1: str, text2: str) -> bool:
        """
        【新增】检查两个文本是否语义等价
        
        Args:
            text1: 文本1
            text2: 文本2
            
        Returns:
            是否语义等价
        """
        # 语义等价词典
        semantic_equivalents = {
            # 维度相关
            '二维': ['2D', '2d', 'two-dimensional', '二维空间'],
            '2D': ['二维', '2d', 'two-dimensional'],
            '2d': ['二维', '2D', 'two-dimensional'],
            '三维': ['3D', '3d', 'three-dimensional', '三维空间'],
            '3D': ['三维', '3d', 'three-dimensional'],
            '3d': ['三维', '3D', 'three-dimensional'],
            
            # 颜色相关
            '彩色': ['彩屏', 'color', 'Color', '全彩'],
            '黑白': ['单色', 'monochrome', 'black and white', '灰度'],
            
            # 触摸相关
            '触摸': ['触摸屏', 'touch', 'touchscreen', '触控'],
            '非触摸': ['普通屏', 'non-touch', '非触控'],
            
            # 无线相关
            '无线': ['wireless', 'WiFi', 'WIFI', 'Wi-Fi', '蓝牙'],
            '有线': ['wired', '有线连接'],
            
            # 单位相关
            '毫米': ['mm', 'millimeter'],
            '厘米': ['cm', 'centimeter'],
            '米': ['m', 'meter'],
            '千克': ['kg', 'kilogram'],
            '克': ['g', 'gram'],
            '升': ['L', 'liter'],
            '毫升': ['mL', 'ml', 'milliliter'],
        }
        
        text1_lower = text1.lower()
        text2_lower = text2.lower()
        
        # 直接相等
        if text1_lower == text2_lower:
            return True
        
        # 检查是否在等价词典中
        for key, equivalents in semantic_equivalents.items():
            if (key.lower() in text1_lower or key.lower() in text2_lower):
                # 检查另一个是否包含等价词
                for equiv in equivalents:
                    if (equiv.lower() in text1_lower and key.lower() in text2_lower) or \
                       (equiv.lower() in text2_lower and key.lower() in text1_lower):
                        return True
        
        return False
    
    def _add_header_instruction(self, output_file: str, total_columns: int):
        """
        【新增】在第一行添加说明信息
        
        Args:
            output_file: 输出Excel文件路径
            total_columns: 总列数
        """
        try:
            from openpyxl.styles import Font
            
            wb = load_workbook(output_file)
            ws = wb.active
            
            # 在第一行插入一行
            ws.insert_rows(1)
            
            # 合并第一行所有列
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
            
            # 设置说明文本
            instruction_text = (
                "颜色说明: 蓝色=供应商数据达标 | 灰色=供应商无数据 | 黄色=需人工审核。 "
                "注意: 该表格数据仅供参考！实际还请人工判断。"
            )
            
            cell = ws.cell(row=1, column=1)
            cell.value = instruction_text
            
            # 设置样式：加粗、四号字体、14磅、黄色背景
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 设置行高
            ws.row_dimensions[1].height = 30
            
            # 冻结前两行（说明行和标题行）
            ws.freeze_panes = 'A3'
            
            wb.save(output_file)
            
        except Exception as e:
            self.log(f"添加说明信息时出错: {str(e)}")
    
    def _apply_yellow_highlight_for_manual_review(self, output_file: str, fusion_types: list):
        """
        为需人工审核的融合数据添加黄色背景
        
        Args:
            output_file: 输出Excel文件路径
            fusion_types: 融合类型列表
        """
        try:
            # 加载workbook
            wb = load_workbook(output_file)
            ws = wb.active
            
            # 创建黄色填充样式
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # 找到"融合数据"列的索引
            # 由于插入了融合数据、融合类型、合并数据三列，融合数据应该在倒数第三列
            # 总列数 = 原供应商列数 + 3个新增列
            fused_data_col_index = ws.max_column - 2  # 融合数据列
            
            # 【修改】遍历fusion_types，为"需人工审核"的行添加黄色背景
            # 由于插入了说明行，数据从第3行开始（第1行是说明，第2行是标题）
            for row_idx, fusion_type in enumerate(fusion_types, start=3):  # 从第3行开始
                if "需人工审核" in str(fusion_type):
                    # 获取该行的融合数据单元格
                    cell = ws.cell(row=row_idx, column=fused_data_col_index)
                    # 应用黄色背景
                    cell.fill = yellow_fill
            
            # 保存workbook
            wb.save(output_file)
            
        except Exception as e:
            self.log(f"添加黄色背景时出错: {str(e)}")
    
    def _apply_vendor_compliance_colors(self, output_file: str, vendor_colors: list, total_rows: int):
        """
        【新增】为供应商列添加达标判定颜色
        
        Args:
            output_file: 输出Excel文件路径
            vendor_colors: 供应商颜色列表 [[color1_row1, color1_row2, ...], [color2_row1, ...]]
            total_rows: 数据总行数
        """
        try:
            wb = load_workbook(output_file)
            ws = wb.active
            
            # 创建颜色填充样式
            blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # 达标→蓝色
            gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # 无数据→灰色
            
            # 供应商列从第2列开始（第1列是参数名称）
            num_vendors = len(vendor_colors)
            
            # 【修改】由于插入了说明行，数据从第3行开始
            for row_idx in range(total_rows):
                for vendor_idx in range(num_vendors):
                    if row_idx < len(vendor_colors[vendor_idx]):
                        color = vendor_colors[vendor_idx][row_idx]
                        cell = ws.cell(row=row_idx + 3, column=vendor_idx + 2)  # +3因为有说明行和标题行
                        
                        # 【修改】达标→蓝色，不达标→不标色
                        if color == 'blue':
                            cell.fill = blue_fill
                        elif color == 'gray':
                            cell.fill = gray_fill
            
            wb.save(output_file)
            
        except Exception as e:
            self.log(f"应用供应商颜色时出错: {str(e)}")    
    
    def _merge_merged_data_column(self, output_file: str, total_rows: int):
        """
        合并"合并数据"列的所有数据单元格
        
        Args:
            output_file: 输出Excel文件路径
            total_rows: 数据总行数
        """
        try:
            # 加载workbook
            wb = load_workbook(output_file)
            ws = wb.active
            
            # 找到"合并数据"列的索引
            # "合并数据"是最后一列
            merged_data_col_index = ws.max_column  # 最后一列
            
            # 【修改】合并拓展范围：由于插入了说明行，从第3行到最后一行
            if total_rows > 0:
                start_row = 3  # 从数据行开始（第1行是说明，第2行是标题）
                end_row = total_rows + 2  # 总行数 + 2(说明行+标题行)
                
                # 执行合并
                merge_range = f"{get_column_letter(merged_data_col_index)}{start_row}:{get_column_letter(merged_data_col_index)}{end_row}"
                ws.merge_cells(merge_range)
            
            # 保存workbook
            wb.save(output_file)
            
        except Exception as e:
            self.log(f"合并合并数据列时出错: {str(e)}")
    
    def _format_cells(self, output_file: str, total_rows: int):
        """
        设置单元格格式：全部自动换行，合并列顶端对齐，其他列居中
        
        Args:
            output_file: 输出Excel文件路径
            total_rows: 数据总行数
        """
        try:
            # 加载workbook
            wb = load_workbook(output_file)
            ws = wb.active
            
            # 获取"合并数据"列的索引（最后一列）
            merged_data_col_index = ws.max_column
            
            # 【修改】遍历所有单元格（从第2行开始，不包括说明行）
            for row in ws.iter_rows(min_row=2, max_row=total_rows + 2, min_col=1, max_col=ws.max_column):
                for cell in row:
                    # 设置自动换行
                    if cell.column == merged_data_col_index:
                        # 合并列：顶端对齐
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                    else:
                        # 其他列：居中对齐
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            
            # 保存workbook
            wb.save(output_file)
            
        except Exception as e:
            self.log(f"设置单元格格式时出错: {str(e)}")
    
    def _save_log(self):
        """保存日志文件"""
        try:
            log_dir = Path('output')
            log_dir.mkdir(exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            log_file = log_dir / f"fusion_log_{timestamp}.txt"
            
            with open(log_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(self.log_messages))
            
            self.log(f"日志已保存到: {log_file}")
            
        except Exception as e:
            print(f"保存日志失败: {str(e)}")


def main():
    """主函数"""
    print("="*60)
    print("医疗设备参数数据智能融合工具")
    print("="*60)
    print()
    
    # 获取输入文件
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        input_file = input("请输入Excel文件路径: ").strip().strip('"')
    
    # 检查文件扩展名
    if not input_file.lower().endswith(('.xlsx', '.xls')):
        print("错误: 请提供Excel文件 (.xlsx 或 .xls)")
        return
    
    # 创建处理器
    processor = MedicalDeviceFusion()
    
    # 处理文件
    success = processor.process_excel(input_file)
    
    if success:
        print("\n" + "="*60)
        print("处理完成!")
        print("="*60)
    else:
        print("\n处理失败，请检查日志信息")
    
    # 等待用户输入
    input("\n按回车键退出...")


if __name__ == '__main__':
    main()
